import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd
import openpyxl
from openpyxl.styles.numbers import FORMAT_PERCENTAGE
import os

pd.set_option('display.max_columns', None)

# cwd = os.getcwd()
# cur_dir = os.chdir("C:/Users/nick-/Desktop/TG_test/")
# print(os.listdir('.'))

'''работаем с дирикторией и файлами в ней'''
os.chdir('C:/Users/nick-/Desktop/TG_test/SH_21_03_2023')
'создаем список файлов'
listOfFiles = list()
for(dirpath, dirnames, filenames) in os.walk(os.getcwd()):
    listOfFiles += [os.path.join(dirpath, file) for file in filenames]
'''делаем шортлист (ЕСЛИ НЕОБХОДИМО'''
shortlist = listOfFiles.copy()
# print(shortlist)

'''обрабатываем выбранные файлы'''
count = 0
for each in shortlist:
    '''получаем название ниши'''
    cur_file_name = each
    cur_file_name = cur_file_name.rstrip('.xlsx')
    cur_file_name = cur_file_name.lstrip(r'C:\Users\nick-\Desktop\TG_test\SH_21_03_2023\Отчет ')
    niche_name = cur_file_name

    '''работаем с репортом по товару'''
    df = pd.read_excel(each)
    df = df.dropna(subset=['Отзывы'])
    df['Отзывы'] = df['Отзывы'].astype(int)

    # print(df.head())
    # print(df.info())

    df_sorted = df.sort_values(by='Продажи, 30 дней (руб)', ascending=False).reset_index()


    sellers_df_sorted = df.pivot_table(index=['Продавец'],
                                    values=['Продажи, 30 дней (руб)', 'Отзывы','Рейтинг','Название'],
                                    aggfunc={'Продажи, 30 дней (руб)': np.sum, 'Отзывы': np.mean, 'Рейтинг': np.mean, 'Название': len}).sort_values(by='Продажи, 30 дней (руб)', ascending=False).reset_index()

    brands_df_sorted = df.pivot_table(index=['Бренд'],
                                    values=['Продажи, 30 дней (руб)', 'Отзывы','Рейтинг','Название'],
                                    aggfunc={'Продажи, 30 дней (руб)': np.sum, 'Отзывы': np.mean, 'Рейтинг': np.mean, 'Название': len}).sort_values(by='Продажи, 30 дней (руб)', ascending=False).reset_index()

    'считаем общую выручку'
    total_revenue = df_sorted['Продажи, 30 дней (руб)'].sum()
    total_missed_revenue = df_sorted['Упущенная выручка за 30 дней'].sum()
    perc_missed_revenue = round(total_missed_revenue/total_revenue, 2)


    ''' ТОВАРЫ '''

    'выручки на карточку товара'
    total_num_goods = len(df_sorted['Продажи, 30 дней (руб)'])
    goods_10_perc = round(total_num_goods*0.1)
    goods_50_perc = round(total_num_goods*0.5)

    rev_per_good_all = round(np.mean(df_sorted['Продажи, 30 дней (руб)'])) # средняя выручка на карточку
    rev_per_good_top100 = round(np.mean(df_sorted.loc[:99, 'Продажи, 30 дней (руб)'])) # средняя выручка на карточку из топ 100
    rev_per_good_10percentile = round(np.mean(df_sorted.loc[:goods_10_perc, 'Продажи, 30 дней (руб)'])) # средняя выручка на карточку из первого 10% персентиля
    rev_per_good_50percentile = round(np.mean(df_sorted.loc[:goods_50_perc, 'Продажи, 30 дней (руб)'])) # средняя выручка на карточку из первого 50% персентиля

    '% выручки товары'
    perc_rev_top100 = round(df_sorted.loc[:99, 'Продажи, 30 дней (руб)'].sum()/total_revenue, 2) # какую долю выручки занимают первые 100 товаров
    perc_rev_10_percentile = round(df_sorted.loc[:goods_10_perc, 'Продажи, 30 дней (руб)'].sum() / total_revenue, 2) # какую долю выручки занимают первые 10% товаров
    perc_rev_50_percentile = round(df_sorted.loc[:goods_50_perc, 'Продажи, 30 дней (руб)'].sum() / total_revenue, 2) # какую долю выручки занимают первые 50% товаров


    ''' ПРОДАВЦЫ '''
    total_num_seller = len(sellers_df_sorted['Продажи, 30 дней (руб)'])
    seller_10_perc = round(total_num_seller*0.1)
    seller_50_perc = round(total_num_seller*0.5)



    '% выручки продавцов'
    seller_perc_rev_top10 = round(sellers_df_sorted.loc[:9, 'Продажи, 30 дней (руб)'].sum()/total_revenue, 2)
    seller_perc_rev_top30 = round(sellers_df_sorted.loc[:29, 'Продажи, 30 дней (руб)'].sum()/total_revenue, 2)
    seller_perc_rev_10_percentile = round(sellers_df_sorted.loc[:seller_10_perc, 'Продажи, 30 дней (руб)'].sum() / total_revenue, 2)
    seller_perc_rev_50_percentile = round(sellers_df_sorted.loc[:seller_50_perc, 'Продажи, 30 дней (руб)'].sum() / total_revenue, 2)
    # print(seller_perc_rev_top10, seller_perc_rev_top30, seller_perc_rev_10_percentile, seller_perc_rev_50_percentile)


    'кол-во товаров на продавца'
    av_goods_per_seller_all = round(np.mean(sellers_df_sorted.loc[:, 'Название']),1) # среднее кол-во товаров на продавца по всем продавцам
    median_goods_pre_seller_all = round(np.median(sellers_df_sorted.loc[:, 'Название']),1) # медианное кол-во товаров на продавца по все продавцам
    av_goods_per_seller_top10perc = round(np.mean(sellers_df_sorted.loc[:seller_10_perc, 'Название']),1) # среднее кол-во товаров на продавца среди 10% самых успешных продавцов

    # print(av_goods_per_seller_all, av_goods_per_seller_top10perc, median_goods_pre_seller_all)


    ''' РЕЙТИНГИ И ОТЗЫВЫ'''

    'отзывы'
    av_reviews_all = round(np.mean(df_sorted.loc[:, 'Отзывы']))
    av_review_top10 = round(np.mean(df_sorted.loc[:9, 'Отзывы']))
    av_review_top30 = round(np.mean(df_sorted.loc[:29, 'Отзывы']))
    av_review_top100 = round(np.mean(df_sorted.loc[:99, 'Отзывы']))

    low_median_goods = round(total_num_goods/2)-20
    upper_median_goods = round(total_num_goods/2)+20
    med_review_all = round(np.median(df_sorted.loc[low_median_goods:upper_median_goods, 'Отзывы'])) # среднее кол-во отзывов для медианного среднеуспешного магазина

    # print(av_reviews_all, av_review_top10, av_review_top30, av_review_top100, med_review_all)

    'рейтинг'

    av_rating_goods_all = round(np.mean(df_sorted.loc[:, 'Рейтинг']), 2)
    av_rating_goods_top100 = round(np.mean(df_sorted.loc[:99, 'Рейтинг']), 2)

    av_rating_seller_top30 = round(np.mean(sellers_df_sorted.loc[:29, 'Рейтинг']), 2)

    # print(av_rating_goods_all, av_rating_goods_top100, av_rating_seller_top30)

    ''' ЦЕНОВЫЕ СЕГМЕНТЫ '''
    av_price_top100 = round(np.mean(df_sorted.loc[:99, 'Цена']))  # средняя цена на карточку в топ100 карточек
    price_perc_25 = round(np.percentile(df_sorted.loc[:, 'Цена'], 25)) # 25 персентиль цены всех карточек
    price_median = round(np.percentile(df_sorted.loc[:, 'Цена'], 50)) # 50 персентиль цены всех карточек
    price_perc_75 = round(np.percentile(df_sorted.loc[:, 'Цена'], 75)) # 75 персентиль цены всех карточек

    # price_stdev = round(np.std(df_sorted.loc[:99, 'Цена со скидкой']))

    # print(av_price_top100, price_perc_25, price_median, price_perc_75)


    ''' ВЕРОЯТНОСТЬ ДОСТИЖЕНИЯ ВЫРУЧКИ'''
    goods_100k_rev = round(len(df_sorted[df_sorted['Продажи, 30 дней (руб)'] >= 100000]) / total_num_goods, 2) # % карточек с выручкой более 100к
    seller_500k_rev = round(len(sellers_df_sorted[sellers_df_sorted['Продажи, 30 дней (руб)'] >= 500000]) / total_num_seller, 2) # % продавцов с выручкой более 500к


    'топ 15 самых дорогих товаров'

    'топ 15 самых дешевых товаров товаров и почему'



    '''подгружаем excel, куда будем записывать данные'''
    wb = openpyxl.load_workbook('C:/Users/nick-/Desktop/TG_test/test_wb_aggregate_SH.xlsx')
    sheet = wb.active
    # print(sheet['A1'].value)
    # print(sheet['A2'].value)

    min_row = sheet.min_row
    max_row = sheet.max_row
    min_col = sheet.min_column
    max_col = sheet.max_column

    # print(min_row, min_col, max_row, max_col)

    ''' записываем значения'''
    row = 1
    col = max_col + 1
    cell = sheet.cell(row = row, column = col)
    cell.value = niche_name

    #общая выручка
    row = row + 1
    cell = sheet.cell(row = row, column = col)
    cell.value = total_revenue
    # упущенная выручка
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = total_missed_revenue
    # % упущенной выручки
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = perc_missed_revenue
    cell.number_format = FORMAT_PERCENTAGE
    # всего карточек в анализе
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = total_num_goods
    # % выручки топ100 товаров
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = perc_rev_top100
    cell.number_format = FORMAT_PERCENTAGE
    # % выручки топ 10% товаров
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = perc_rev_10_percentile
    cell.number_format = FORMAT_PERCENTAGE
    # % выручки топ 50% товаров
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = perc_rev_50_percentile
    cell.number_format = FORMAT_PERCENTAGE
    # % карточек с более чем 100т выручкой
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = goods_100k_rev
    cell.number_format = FORMAT_PERCENTAGE
    # Выручка на 1 карточку товара (все)
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = rev_per_good_all
    # Выручка на 1 карточку товара (топ100)
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = rev_per_good_top100
    # Выручка на 1 карточку товара (топ 10%)
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = rev_per_good_10percentile
    # Выручка на 1 карточку товара (топ 50%)
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = rev_per_good_50percentile
    # всего продавцов
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = total_num_seller
    # % продавцов с выручкой более 500к
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = seller_500k_rev
    cell.number_format = FORMAT_PERCENTAGE
    # % выручки топ10 продавцов
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = seller_perc_rev_top10
    cell.number_format = FORMAT_PERCENTAGE
    # % выручки топ30 продавцов
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = seller_perc_rev_top30
    cell.number_format = FORMAT_PERCENTAGE
    # % выручки топ 10% продавцов
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = seller_perc_rev_10_percentile
    cell.number_format = FORMAT_PERCENTAGE
    # % выручки топ 50% продавцов
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = seller_perc_rev_50_percentile
    cell.number_format = FORMAT_PERCENTAGE
    # Среднее кол-во товаров на продавца
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = av_goods_per_seller_all
    # Медианное кол-во товаров на продавца
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = median_goods_pre_seller_all
    # Среднее кол-во товаров у топ 10% продавцов
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = av_goods_per_seller_top10perc
    # Среднее кол-во отзывов
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = av_reviews_all
    # Кол-во отзывов у топ-10
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = av_review_top10
    # Кол-во отзывов у топ-30
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = av_review_top30
    # Кол-во отзывов у топ-100
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = av_review_top100
    # Кол-во отзывов у медианного товара
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = med_review_all
    # Средний рейтинг
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = av_rating_goods_all
    # Средний рейтинг у топ 30 товаров
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = av_rating_seller_top30
    # Средний рейтинг у топ 100 товаров
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = av_rating_goods_top100
    # Средняя цена (топ 100 карточек)
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = av_price_top100
    # 25 персентиль цены
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = price_perc_25
    # Медианная цена
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = price_median
    # 75 персентиль цены
    row = row + 1
    cell = sheet.cell(row=row, column=col)
    cell.value = price_perc_75



    wb.save('C:/Users/nick-/Desktop/TG_test/test_wb_aggregate_SH.xlsx')

    count += 1
    print(count)